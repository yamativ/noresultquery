############################################################\
### weekly no_result_query one stop query\
### https://quip-apple.com/XszzAdZFVofr\
### last update 20230727\
############################################################\
\
\
source myenv/bin/activate # activating the virual environment \
python3                   # activating python3\
\
\
############################################################\
################ Bilding weekly table ######################\
############################################################\
\
# library import \
import subprocess # for runnnig hive and bash query\
import datetime   # to get the target week\
import sys        # for closing python\
\
### determing table names: \
# output_table name will be a date of monday, 1 week before the code executed.\
# other table names are used for calling previouse 2,3,4 weeks table for table join.\
output_table = (datetime.datetime.now() - datetime.timedelta(days=datetime.datetime.now().weekday() - 1 + 8)).strftime('%Y_%m_%d')\
w2_before_table = (datetime.datetime.now() - datetime.timedelta(days=datetime.datetime.now().weekday() - 1 + 15)).strftime('%Y_%m_%d')\
w3_before_table = (datetime.datetime.now() - datetime.timedelta(days=datetime.datetime.now().weekday() - 1 + 22)).strftime('%Y_%m_%d')\
w4_before_table = (datetime.datetime.now() - datetime.timedelta(days=datetime.datetime.now().weekday() - 1 + 29)).strftime('%Y_%m_%d')\
\
### defining hive query for controling hive from python\
# creating weekly table and joined past 4 weeks tables\
table_create = f'''\
USE no_result_query;\
MSCK REPAIR TABLE auto_complete_unified;\
SET hive.exec.dynamic.partition.mode=nonstrict;\
SET hivevar:log_db=no_result_query;\
SET hivevar:log_table=auto_coamplete_unified;\
\
-- create blank latest weekly table --\
-- DROP TABLE no_result_query.\{output_table\}\
CREATE TABLE IF NOT EXISTS no_result_query.\{output_table\}(\
  query STRING,\
  search_count INT,\
  ac_no_result_count INT,\
  search_no_result_count INT,\
  ac_ratio DOUBLE,\
  seach_ratio DOUBLE\
)\
PARTITIONED BY (date_from DATE, date_to DATE)\
ROW FORMAT SERDE 'org.apache.hadoop.hive.serde2.OpenCSVSerde'\
WITH SERDEPROPERTIES (\
  "separatorChar" = "\\t",\
  "quoteChar" = '\\\\"',\
  "escapeChar" = '\\\\\\\\'\
);\
\
-- insert data into latest weekly table --\
INSERT OVERWRITE TABLE no_result_query.\{output_table\} \
PARTITION (\
  date_from,\
  date_to\
) \
SELECT\
  query,\
  COUNT(*) AS search_count,\
  SUM(is_ac_no_result) AS ac_no_result_count,\
  SUM(is_search_no_result) AS search_no_result_count,\
  SUM(is_ac_no_result) / COUNT(*)  AS ac_ratio,\
  SUM(is_search_no_result) / COUNT(*) seach_ratio,\
  to_date(date_sub(current_date(), cast(date_format(current_date(), 'u') as int) - 1 + 7)) AS date_from,\
  date_add(to_date(date_sub(current_date(), cast(date_format(current_date(), 'u') as int) - 1 + 7)), 6) AS date_to\
FROM (\
    SELECT\
      regexp_replace(acquery, '\\t', ' ') AS query,\
      CASE WHEN autoCompleteResultsCount = 0 THEN 1 ELSE 0 END AS is_ac_no_result,\
      CASE WHEN searchResultCount = 0 THEN 1 ELSE 0 END AS is_search_no_result\
    FROM no_result_query.auto_complete_unified\
    WHERE appId = 'com.apple.Maps'\
    AND locale IN ('ja_JP', 'ja-JP_JP')\
    AND hardwareModel LIKE 'iPhone%'\
    AND ds >= to_date(date_sub(current_date(), cast(date_format(current_date(), 'u') as int) - 1 + 7))\
    AND ds <= date_add(to_date(date_sub(current_date(), cast(date_format(current_date(), 'u') as int) - 1 + 7)), 6)\
    AND acquery != ''\
    AND actionType = 'CLICK_SEARCH'\
) t\
GROUP BY query\
HAVING search_count >= 10\
ORDER BY search_count DESC\
;\
\
\
-- merge past 3 week data to latest weekly table --\
-- DROP TABLE no_result_query.recent_4w_\{output_table\}\
CREATE EXTERNAL TABLE IF NOT EXISTS no_result_query.recent_4w_\{output_table\}(\
  query STRING,\
  total_count BIGINT,\
  search_count_w4 INT, ac_no_result_count_w4 INT, search_no_result_count_w4 INT, ac_ratio_w4 DOUBLE, seach_ratio_w4 DOUBLE,\
  search_count_w3 INT, ac_no_result_count_w3 INT, search_no_result_count_w3 INT, ac_ratio_w3 DOUBLE, seach_ratio_w3 DOUBLE,\
  search_count_w2 INT, ac_no_result_count_w2 INT, search_no_result_count_w2 INT, ac_ratio_w2 DOUBLE, seach_ratio_w2 DOUBLE,\
  search_count_w1 INT, ac_no_result_count_w1 INT, search_no_result_count_w1 INT, ac_ratio_w1 DOUBLE, seach_ratio_w1 DOUBLE,\
  max_ac_sc DOUBLE,\
  trend FLOAT\
)\
ROW FORMAT SERDE 'org.apache.hadoop.hive.serde2.OpenCSVSerde'\
WITH SERDEPROPERTIES (\
  "separatorChar" = "\\t",\
  "quoteChar" = '\\\\"',\
  "escapeChar" = '\\\\\\\\'\
);\
\
-- insert data into merged table --\
INSERT OVERWRITE TABLE no_result_query.recent_4w_\{output_table\}\
SELECT \
query,\
search_count_w1+search_count_w2+search_count_w3+search_count_w4 AS total_count,\
search_count_w4,ac_no_result_count_w4,search_no_result_count_w4,ac_ratio_w4,seach_ratio_w4,\
search_count_w3,ac_no_result_count_w3,search_no_result_count_w3,ac_ratio_w3,seach_ratio_w3,\
search_count_w2,ac_no_result_count_w2,search_no_result_count_w2,ac_ratio_w2,seach_ratio_w2,\
search_count_w1,ac_no_result_count_w1,search_no_result_count_w1,ac_ratio_w1,seach_ratio_w1,\
(CASE WHEN ac_ratio_w1 > seach_ratio_w1 THEN ac_ratio_w1 \
WHEN ac_ratio_w1 < seach_ratio_w1 THEN seach_ratio_w1\
ELSE 0 END) AS max_ac_sc,\
search_count_w1 / (\
(CASE WHEN  search_count_w4 = 0 THEN 7.5 ELSE search_count_w4 END +\
CASE WHEN  search_count_w3 = 0 THEN 7.5  ELSE search_count_w3 END +\
CASE WHEN  search_count_w2 = 0 THEN 7.5  ELSE search_count_w2 END) / 3\
) AS trend\
FROM (\
SELECT\
w1.query, \
nvl(w4.search_count, 0) AS search_count_w4, w4.ac_no_result_count AS ac_no_result_count_w4, w4.search_no_result_count AS search_no_result_count_w4, w4.ac_ratio AS ac_ratio_w4, w4.seach_ratio AS seach_ratio_w4,\
nvl(w3.search_count, 0) AS search_count_w3, w3.ac_no_result_count AS ac_no_result_count_w3, w3.search_no_result_count AS search_no_result_count_w3, w3.ac_ratio AS ac_ratio_w3, w3.seach_ratio AS seach_ratio_w3,\
nvl(w2.search_count, 0) AS search_count_w2, w2.ac_no_result_count AS ac_no_result_count_w2, w2.search_no_result_count AS search_no_result_count_w2, w2.ac_ratio AS ac_ratio_w2, w2.seach_ratio AS seach_ratio_w2,\
nvl(w1.search_count, 0) AS search_count_w1, w1.ac_no_result_count AS ac_no_result_count_w1, w1.search_no_result_count AS search_no_result_count_w1, w1.ac_ratio AS ac_ratio_w1, w1.seach_ratio AS seach_ratio_w1\
FROM no_result_query.\{output_table\} w1\
LEFT JOIN no_result_query.\{w2_before_table\} w2 ON w1.query = w2.query\
LEFT JOIN no_result_query.\{w3_before_table\} w3 ON w1.query = w3.query\
LEFT JOIN no_result_query.\{w4_before_table\} w4 ON w1.query = w4.query\
)t\
;'''\
\
# excuting the hive query defined ave\
subprocess.run(['hive', '-e', table_create])\
\
# defining the bash query for copying the table to serevr/home/user\
to_user_server = f'hive -e"select * from no_result_query.recent_4w_\{output_table\}" | head -n -2 > recent_4w_\{output_table\}'\
subprocess.run(['bash','-c', to_user_server]) # excute above\
\
\
### Activate for separate run\
#source myenv/bin/activate # activating the virual environment \
#python3                   # activating python3\
\
############################################################\
################ Convert to Excel FIle #####################\
############################################################\
\
# library import \
import subprocess # for runnnig hive and bash query\
import datetime   # to get the target week\
import sys        # for closing python\
import re\
import pandas as pd \
import urllib.parse\
from openpyxl import workbook, load_workbook\
from openpyxl.styles import Alignment,Border, Side, Font, PatternFill\
from openpyxl.utils.dataframe import dataframe_to_rows\
from openpyxl.worksheet.datavalidation import DataValidation\
\
\
hvfile = 'recent_4w_' + output_table # set file name created in previouse section\
exfile = hvfile + '.xlsx'           # Set excel file name\
\
# Read TSV file, the marged weekly table\
df = pd.read_csv(hvfile, delimiter='\\t', header=None)\
\
\
### Add Charachter columns\
# Set udf\
def detect_character_type(text):\
    hiragana_pattern = r"[\\u3041-\\u3096\uc0\u12540 ]"\
    katakana_pattern = r"[\\u30A1-\\u30FA\uc0\u12540 ]"\
    kanji_pattern = r"[\\u4E00-\\u9FFF]"\
    alphabet_pattern = r"[a-zA-Z]"\
    number_pattern = r"\\d"  # \uc0\u25968 \u23383 \u12398 \u12497 \u12479 \u12540 \u12531 \
    \
    character_types = []\
    \
    if re.search(hiragana_pattern, text):\
        character_types.append("\uc0\u12402 \u12425 \u12364 \u12394 ")\
    if re.search(katakana_pattern, text):\
        character_types.append("\uc0\u12459 \u12479 \u12459 \u12490 ")\
    if re.search(kanji_pattern, text):\
        character_types.append("\uc0\u28450 \u23383 ")\
    if re.search(alphabet_pattern, text):\
        character_types.append("\uc0\u12450 \u12523 \u12501 \u12449 \u12505 \u12483 \u12488 ")\
    if re.search(number_pattern, text):\
        character_types.append("\uc0\u25968 \u23383 ")  # \u25968 \u23383 \u12398 \u21028 \u23450 \
    unique_character_types = list(set(character_types))  # \uc0\u37325 \u35079 \u12434 \u38500 \u21435 \
    return "\uc0\u12539 ".join(unique_character_types)\
\
character = []\
for i in df.iloc[:,0]:\
    character_type = detect_character_type(i)\
    character.append(character_type)\
\
character_df = pd.DataFrame(character)\
#character_df.name = 'character'\
df = pd.concat([df.iloc[:, 0], character_df, df.iloc[:, 1:]], axis=1)\
\
\
### Add Columns Name and Filtering/Sorting\
# Define adn Add Colmuns Name\
df.columns = ['query','character','total_count',\
              'search_count_w4','ac_no_result_count_w4','search_no_result_count_w4','ac_ratio_w4','seach_ratio_w4',\
              'search_count_w3','ac_no_result_count_w3','search_no_result_count_w3','ac_ratio_w3','seach_ratio_w3',\
              'search_count_w2','ac_no_result_count_w2','search_no_result_count_w2','ac_ratio_w2','seach_ratio_w2',\
              'search_count_w1','ac_no_result_count_w1','search_no_result_count_w1','ac_ratio_w1','seach_ratio_w1',\
              'max_ac_sc','trend']  \
\
# Sort colmuns and get top 1000\
df_sorted = df[df['max_ac_sc']>= 0.5]\
df_sorted = df_sorted.sort_values(['search_count_w1','max_ac_sc','trend','total_count'], ascending=[False,False,False,False])\
df_sorted_1k = df_sorted.head(1000)\
\
# Set the target columns for converting to decimal 3\
columns_convert = ['ac_ratio_w4','seach_ratio_w4',\
                   'ac_ratio_w3','seach_ratio_w3',\
                   'ac_ratio_w2','seach_ratio_w2',\
                   'ac_ratio_w1','seach_ratio_w1',\
                   'max_ac_sc','trend']\
\
# Convet decimal 3\
df_sorted_1k[columns_convert] = round(df_sorted_1k[columns_convert],3)\
\
# Save data as a Excel file as "recent_4w_yyyy_mm_dd.xlsx"\
df_sorted_1k.to_excel(exfile, index=False)\
\
\
### Excel Customize\
# Open Previouse Excel File \
wb = load_workbook(exfile)\
\
# Activate the target work sheet\
ws1 = wb.active\
\
### Setting \
# Additional Columns Name List\
column_names = ['Classification', 'Status', 'Issue Type', 'In Prod', 'Comment', 'Gmap', 'Gsearch', 'AGSearch', 'Amap', 'Apollo', 'Hawk']\
\
# set width for columnA\
ws1.column_dimensions['A'].width = 40\
\
# Cell hight\
ws1.row_dimensions[1].height = 21\
\
# Ruled Lines \
border = Border(\
    left=Side(border_style='thin'),\
    right=Side(border_style='thin'),\
    top=Side(border_style='thin'),\
    bottom=Side(border_style='thin')\
)\
\
# Add columns name from Additional Columns Name list\
for col_num, column_name in enumerate(column_names, start=26):  # Add from row 26 = 'Z'\
    cell = ws1.cell(row=1, column=col_num)\
    cell.value = column_name\
    cell.alignment = Alignment(horizontal='center', vertical='center')\
    cell.border = border\
    cell.font = Font(bold=True)\
\
# Headder Cell / Font Color \
for i in range(1,26):\
    cell = ws1.cell(row=1, column=i)\
    cell.fill = PatternFill(fill_type="solid", fgColor="D9D9D9")\
\
for i in range(26,31):\
    cell = ws1.cell(row=1, column=i)\
    cell.fill = PatternFill(fill_type="solid", fgColor="808080")\
    cell.font = Font(color="FFFFFF")\
\
# Higlith latest weeks query count    \
for i in range(1, 1001):\
    cell = ws1.cell(row=i, column=19)\
    cell.fill = PatternFill(fill_type="solid", fgColor="FCD5B4")\
\
# Highlight new query in 4 weeks\
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):\
    c_value = row[3].value if row[3].value is not None else 0 # C row value\
    h_value = row[8].value if row[8].value is not None else 0 # H row value\
    m_value = row[15].value if row[15].value is not None else 0 # M row value\
    \
    if c_value + h_value + m_value == 0:\
        cell = row[0] # cell in row A \
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")\
        cell.font = Font(color="000000")  # black\
\
# Customize columns width\
col_wdth = ['B','C','H','M','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']\
for i in col_wdth:\
    ws1.column_dimensions[i].width = 15\
\
# Freeze Row 1 and Column A\
ws1.freeze_panes = 'B2'  \
\
# Add filter for all comuns\
ws1.auto_filter.ref = ws1.dimensions\
\
# Hide Colmuns B to S\
ws1.column_dimensions.group('E', 'H', hidden=True)\
ws1.column_dimensions.group('J', 'M', hidden=True)\
ws1.column_dimensions.group('O', 'R', hidden=True)\
\
\
### Add Dropdown List\
# Define List Contents\
classification_list = ["POI", "Brand", "Category", "Noise", "Function", "Event",\
                       "Basemap Feature", "Natural Feature", "Building", "Address", \
                       "Geocode", "Territory", "Transit", "Foreign POI", "Foreign Basemap"]\
\
status_list = ["Unnecessary", "Good", "BinB", "Partial", "No Result", "NG", "Bad"]\
\
issuetype_list = ["Not in Gemini", "Renamed", "Inactive", "Alias", "Limited Data",\
                  "Category Trigger", "Unclassifiable Category","No JP Name","Mixed intent",\
                  "Pending", "Unnecessary","Foreign Query","Search Logic"]\
\
inprod_list = ["Unnecessary", "TRUE", "FALSE", "Minor Issue","In Gemini", "AC"]\
\
# Define the cell range that the list utilized\
classification_range = 'Z2:Z1001'\
status_range = 'AA2:AA1001'\
issuetype_range = 'AB2:AB1001'\
inprod_range = 'AC2:AC1001'\
\
# Drop Down List Setting\
classification_dv = DataValidation(type='list',formula1=f'"\{",".join(classification_list)\}"')\
status_dv = DataValidation(type='list',formula1=f'"\{",".join(status_list)\}"')\
issuetype_dv = DataValidation(type='list',formula1=f'"\{",".join(issuetype_list)\}"')\
inprod_dv = DataValidation(type='list',formula1=f'"\{",".join(inprod_list)\}"')\
\
# Utilize dv to object cell range\
ws1.add_data_validation(classification_dv)\
classification_dv.add(classification_range)\
\
ws1.add_data_validation(status_dv)\
status_dv.add(status_range)\
\
ws1.add_data_validation(issuetype_dv)\
issuetype_dv.add(issuetype_range)\
\
ws1.add_data_validation(inprod_dv)\
inprod_dv.add(inprod_range)\
\
\
# Write List in Sheet2\
ws2 = wb.create_sheet(title='Drop down list')\
\
# Imput List data in to sheet \
for i,value in enumerate(classification_list, start = 1):\
    ws2[f'A\{i\}'] = value\
\
for i,value in enumerate(status_list, start = 1):\
    ws2[f'B\{i\}'] = value    \
\
for i,value in enumerate(issuetype_list, start = 1):\
    ws2[f'C\{i\}'] = value    \
\
for i,value in enumerate(inprod_list, start = 1):\
    ws2[f'D\{i\}'] = value\
\
    \
## Add Hyper Link\
# hyper link of google map, google search, google search by term, apple maps, apollo, hawk \
# preparation for week_url\
\
min_date = (datetime.datetime.now() - datetime.timedelta(days=datetime.datetime.now().weekday() - 1 + 8)).strftime('%Y%m%d')\
formatted_min_date = f"\{min_date[4:6]\}%2F\{min_date[6:8]\}%2F\{min_date[0:4]\}"\
encoded_min_date = f"min%3A\{formatted_min_date\}%2Ccd"\
\
max_date = (datetime.datetime.now() - datetime.timedelta(days=datetime.datetime.now().weekday() - 1 + 2)).strftime('%Y%m%d')\
formatted_max_date = f"\{max_date[4:6]\}%2F\{max_date[6:8]\}%2F\{max_date[0:4]\}"\
encoded_max_date = f"_max%3A\{formatted_max_date\}&tbm="\
\
for i in range(2, 1001):\
    value = ws1[f'A\{i\}'].value\
    if value is not None:\
        gmap_url = "https://www.google.com/maps/place/" + ws1[f'A\{i\}'].value\
        ws1[f'AE\{i\}'].hyperlink = gmap_url\
        ws1[f'AE\{i\}'].value = "Gmap\uc0\u12391 \u26908 \u32034 "\
        \
        query_url = "http://www.google.co.jp/search?num=50&q=" + ws1[f'A\{i\}'].value\
        ws1[f'AF\{i\}'].hyperlink = query_url\
        ws1[f'AF\{i\}'].value = "Query\uc0\u12434 \u26908 \u32034 "\
        \
        week_url = "https://www.google.co.jp/search?q=" +  ws1[f'A\{i\}'].value + "&num=50&source=lnt&tbs=cdr%3A1%2Ccd_" + encoded_min_date + encoded_max_date\
        ws1[f'AG\{i\}'].hyperlink = week_url\
        ws1[f'AG\{i\}'].value = "\uc0\u26399 \u38291 \u38480 \u23450 \u26908 \u32034 "\
        \
        amap_url = "https://maps.apple.com/?address=" + ws1[f'A\{i\}'].value\
        ws1[f'AH\{i\}'].hyperlink = amap_url \
        ws1[f'AH\{i\}'].value = "Amap\uc0\u12391 \u26908 \u32034 "\
        \
        appolo_url = "https://apollo.geo.apple.com/?query=" + ws1[f'A\{i\}'].value\
        ws1[f'AI\{i\}'].hyperlink = appolo_url\
        ws1[f'AI\{i\}'].value = "Apollo\uc0\u12391 \u26908 \u32034 "\
        \
        hwk_url = "https://hawk.geo.apple.com/search?center=35.681042, 139.767214&zoom=11&rpc_locale=ja_JP&settingsLanguage=ja_JP&device_country_code=JPN&language=ja-JP&xray&responseView=xray&query=" + ws1[f'A\{i\}'].value\
        ws1[f'AJ\{i\}'].hyperlink = hwk_url\
        ws1[f'AJ\{i\}'].value = "Hawk\uc0\u12391 \u26908 \u32034 "\
\
\
# Save Changes\
wb.save(exfile)\
\
### Delete copyfile from user space\
#delete_row_file = f'rm recent_4w_\{output_table\}' # define bash query \
#subprocess.run(['bash','-c', delete_row_file])   # excute\
\
# escape from python\
sys.exit()\
